import requests, bs4, time, os
from openpyxl import Workbook
from tqdm import tqdm

def main():
    header()    
    researchesList = getResearchesList()
    WorkSheetXLSX(researchesList).sheet()
    print("\nFeito!\n")
    os.system("PAUSE")

def header():
    print("="*50)
    print("   GERADOR DE RELATÓRIO DE PESQUISA ACADÊMICA")
    print("="*50)
    print()
    print("@autor: Robson L. Lopes")
    print("GitHub: https://github.com/Robsonmxms\n")
        
def getResearchesList():
    researchesList = list()

    choiceValue = '1'
    
    while choiceValue == '1':
        choiceValue = getChoiceValue()
        research = getResearchList(choiceValue)
        if research:
            researchesList.append(research)

    return researchesList

def getChoiceValue():
    time.sleep(0.5)
    print("-"*50)
    print("\nDeseja fazer uma nova busca?")
    print("[0] NÃO")
    print("[1] SIM\n")
    choiceValue = input()
    print()
    print("-"*50)

    return choiceValue

def getResearchList(choiceValue):
    researchList = list()

    if choiceValue == '1':
        researchList = Research.research()
    elif choiceValue == '0':
        print("Pesquisas realizadas!\n")            
    else:
        print("ERROR\n")

    return researchList

class Research:
    def __init__(self, title, refUrlsList):
        self.title = title
        self.refUrlsList = refUrlsList

    def getRefUrlsList(self):
        return self.title

    def getUrl(self):
        return self.refUrlsList

    def setTitle(self,title):
        self.title = title

    def setRefUrlsList(self,refUrlsList):
        self.refUrlsList = refUrlsList
    
    def research():
        scholarResearch = ScholarResearch.getScholarResearch()
        refUrlsList = scholarResearch.getRefUrlsList()
        research = Research(scholarResearch.title,refUrlsList)

        print("\nPesquisa concluída!")

        return research

class ScholarResearch:
    def __init__(self, title, url):
        self.title = title
        self.url = url

    def getTitle(self):
        return self.title

    def getUrl(self):
        return self.url

    def setTitle(self,title):
        self.title = title

    def setUrl(self,url):
        self.url = url
        
    def getScholarResearch():
        print("\nDigite algo a ser pesquisado no Google Acadêmico: ")
        title = input()
        searchPhrase = title + " filetype = pdf"

        print("\nPesquisando " + title + "...")

        scholar = 'https://scholar.google.com.br/scholar?hl=pt-BR&as_sdt=0%2C5&q='
        url = scholar + searchPhrase
        scholarResearch = ScholarResearch(title, url)

        return scholarResearch

    def getRefUrlsList(self):
        refUrlsList = self.__getConcatenateRefUrlsList()
        refUrlsList = list(set(refUrlsList))

        return refUrlsList

    def __getConcatenateRefUrlsList(self):

        def isGoogleSearch(refUrl):
            return refUrl[:4] != "http"
    
        concatenateRefUrlsList = self.__getNoConcatenateRefUrlsList()
        scholarUrl = 'https://scholar.google.com.br'

        for i in range(len(concatenateRefUrlsList)):
            if isGoogleSearch(concatenateRefUrlsList[i]):
                concatenateRefUrlsList[i] = scholarUrl + concatenateRefUrlsList[i]

        return concatenateRefUrlsList


    def __getNoConcatenateRefUrlsList(self):
        urlsList = self.__getSoup().find_all("a")

        noConcatenateRefUrlsList = list()

        for i in range(len(urlsList)):
            noConcatenateRefUrlsList.append(urlsList[i].get("href"))

        return noConcatenateRefUrlsList

    
    def __getSoup(self):
        res = requests.get(self.url)

        #verifica erros, interrompendo a execução caso ocorra problemas
        res.raise_for_status()

        soup = bs4.BeautifulSoup(res.text, features="lxml")

        return soup


class WorkSheetXLSX:
    def __init__(self,researchesList):
        self.researchesList = researchesList
        
    def getResearchesList(self):
            return self.researchesList
    
    def setTitle(self,researchesList):
            self.researchesList = researchesList
        
    def sheet(self):
        global ws
        workSheet = Workbook()
        ws = workSheet.active
        ws.title = 'Pesquisas'
        self.__getTable()
        workSheet.save("relatorio.xlsx")

    def __getTable(self):
        pbar = tqdm(range(1,len(self.researchesList)+1))
        for indexColumn in pbar:
            self.__tabulating(pbar,indexColumn)

    def __tabulating(self,pbar,indexColumn):
        research = self.researchesList[indexColumn-1]
        size = len(research.refUrlsList)
        Column(indexColumn,size,research).rows()
        time.sleep(0.5)
        pbar.set_description("Gerando relatório")

class Column:
    def __init__(self, indexColumn, size, research):
        self.indexColumn = indexColumn
        self.size = size
        self.research = research
        
    def getIndexColumn(self):
            return self.indexColumn

    def getSize(self):
            return self.size

    def getResearch(self):
            return self.research
    
    def setSize(self,size):
            self.size = size

    def setResearch(self,research):
            self.research = research

    def setIndexColumn(self,indexColumn):
            self.indexColumn = indexColumn
        
    def rows(self):
        for row in range(1,self.size+1):
            aColumn = Column(
                self.indexColumn,
                self.size,
                self.research
            )
            Cell(row,aColumn).cell()

class Cell:
    def __init__(self, row, aColumn):
        self.row = row
        self.aColumn = aColumn
        
    def getRow(self):
            return self.row

    def getAColumn(self):
            return self.aColumn
    
    def setRow(self,row):
            self.row = row

    def setAColumn(self,aColumn):
            self.aColumn = aColumn
            
    def cell(self):
        
        if self.row == 1:
            self.__doIfMenuHeader()
        else:
            self.__doIfNotMenuHeader()

    def __doIfMenuHeader(self):
        ws.cell(
            row = self.row,
            column = self.aColumn.indexColumn,
            value = self.aColumn.research.title
        )

    def __doIfNotMenuHeader(self):
        url = self.aColumn.research.refUrlsList[self.row-1] 
        ws.cell(
            row = self.row,
            column = self.aColumn.indexColumn,
            value = self.aColumn.research
            .title+'_link'+str(self.row-1)
        ).hyperlink = url

if __name__ == "__main__":
    main()
